from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required
from app import db
from app.models import Partido, Serie, Equipo
from datetime import datetime, date, time
import io

carga_masiva_bp = Blueprint('carga_masiva', __name__)


def _generar_plantilla():
    """Genera el archivo Excel de plantilla para descarga."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # ── Hoja 1: Plantilla de carga ─────────────────────────
    ws = wb.active
    ws.title = "partidos"

    # Estilos
    verde       = "1a7a3c"
    verde_claro = "e8f5e9"
    amarillo    = "e8c547"
    gris_header = "1e1e1e"
    blanco      = "FFFFFF"
    rojo_claro  = "fdecea"

    font_header  = Font(bold=True, color=blanco, size=11)
    fill_header  = PatternFill("solid", fgColor=verde)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Encabezados
    columnas = [
        ("fecha",        "Fecha\n(DD/MM/AAAA)", 16),
        ("hora",         "Hora\n(HH:MM)",        12),
        ("serie_codigo", "Código\nSerie",         14),
        ("local_codigo", "Código\nLocal",         14),
        ("visita_codigo","Código\nVisita",         14),
        ("cancha",       "Cancha\n(opcional)",    20),
        ("arbitro",      "Árbitro\n(opcional)",   20),
    ]

    for col_idx, (_, header, width) in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 36

    # Filas de ejemplo
    ejemplos = [
        ("15/05/2025", "10:00", "SUB10", "DN", "RA", "Cancha Central", "J. Pérez"),
        ("15/05/2025", "11:30", "SUB10", "LV", "CF", "Cancha Norte", ""),
        ("22/05/2025", "09:00", "ADU",   "DN", "LV", "", ""),
    ]

    fill_par  = PatternFill("solid", fgColor="F9F9F9")
    fill_impar = PatternFill("solid", fgColor=blanco)

    for row_idx, fila in enumerate(ejemplos, start=2):
        fill = fill_par if row_idx % 2 == 0 else fill_impar
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = align_left
            cell.fill = fill
            cell.border = border

    # Filas vacías para rellenar (hasta fila 102 = 100 partidos)
    for row_idx in range(len(ejemplos) + 2, 103):
        fill = fill_par if row_idx % 2 == 0 else fill_impar
        for col_idx in range(1, len(columnas) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = align_left

    ws.freeze_panes = "A2"

    # ── Hoja 2: Referencia de códigos ──────────────────────
    ws2 = wb.create_sheet("referencia")
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 30

    # Títulos secciones
    for col, titulo in [(1, "CÓDIGO EQUIPO"), (2, "NOMBRE EQUIPO"),
                        (3, "CÓDIGO SERIE"), (4, "NOMBRE SERIE")]:
        c = ws2.cell(row=1, column=col, value=titulo)
        c.font = Font(bold=True, color=blanco, size=10)
        c.fill = PatternFill("solid", fgColor=verde)
        c.alignment = align_center
        c.border = border

    ws2.cell(row=2, column=1, value="← Los equipos y series")
    ws2.cell(row=3, column=1, value="   se cargan desde la")
    ws2.cell(row=4, column=1, value="   base de datos al")
    ws2.cell(row=5, column=1, value="   descargar la plantilla")

    # Nota
    nota = ws2.cell(row=7, column=1,
        value="⚠️  Usa exactamente los códigos que aparecen aquí.")
    nota.font = Font(bold=True, color="C0392B", size=10)
    ws2.merge_cells("A7:D7")

    wb_bytes = io.BytesIO()
    wb.save(wb_bytes)
    wb_bytes.seek(0)
    return wb_bytes


def _generar_plantilla_con_datos():
    """Genera plantilla con los códigos reales de la BD en hoja de referencia."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "partidos"

    verde      = "1a7a3c"
    blanco     = "FFFFFF"
    font_h     = Font(bold=True, color=blanco, size=11)
    fill_h     = PatternFill("solid", fgColor=verde)
    align_c    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_l    = Alignment(horizontal="left", vertical="center")
    border     = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    fill_par   = PatternFill("solid", fgColor="F5FAF7")
    fill_impar = PatternFill("solid", fgColor=blanco)

    columnas = [
        ("fecha",         "Fecha\n(DD/MM/AAAA)", 16),
        ("hora",          "Hora\n(HH:MM)",        12),
        ("serie_codigo",  "Código\nSerie",         14),
        ("local_codigo",  "Código\nLocal",         14),
        ("visita_codigo", "Código\nVisita",         14),
        ("cancha",        "Cancha\n(opcional)",    20),
        ("arbitro",       "Árbitro\n(opcional)",   20),
    ]

    for col_idx, (_, header, width) in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font_h
        cell.fill = fill_h
        cell.alignment = align_c
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 36

    for row_idx in range(2, 202):
        fill = fill_par if row_idx % 2 == 0 else fill_impar
        for col_idx in range(1, len(columnas) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = align_l

    ws.freeze_panes = "A2"

    # ── Hoja de referencia con datos reales ────────────────
    ws2 = wb.create_sheet("referencia")
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 30

    for col, titulo in [(1, "CÓDIGO EQUIPO"), (2, "NOMBRE EQUIPO"),
                        (3, "CÓDIGO SERIE"),  (4, "NOMBRE SERIE")]:
        c = ws2.cell(row=1, column=col, value=titulo)
        c.font = Font(bold=True, color=blanco, size=10)
        c.fill = PatternFill("solid", fgColor=verde)
        c.alignment = align_c
        c.border = border

    equipos = Equipo.query.filter_by(activo=True).order_by(Equipo.codigo).all()
    series  = Serie.query.filter_by(activa=True).order_by(Serie.codigo).all()

    max_rows = max(len(equipos), len(series))
    fill_e   = PatternFill("solid", fgColor="E8F5E9")
    fill_s   = PatternFill("solid", fgColor="FFF8E1")

    for i in range(max_rows):
        row = i + 2
        if i < len(equipos):
            c1 = ws2.cell(row=row, column=1, value=equipos[i].codigo)
            c2 = ws2.cell(row=row, column=2, value=equipos[i].nombre)
            c1.fill = c2.fill = fill_e
            c1.border = c2.border = border
        if i < len(series):
            c3 = ws2.cell(row=row, column=3, value=series[i].codigo)
            c4 = ws2.cell(row=row, column=4, value=series[i].nombre)
            c3.fill = c4.fill = fill_s
            c3.border = c4.border = border

    nota = ws2.cell(row=max_rows + 3, column=1,
        value="⚠️  Usa exactamente los códigos de esta hoja en la pestaña 'partidos'.")
    nota.font = Font(bold=True, color="C0392B")
    ws2.merge_cells(f"A{max_rows+3}:D{max_rows+3}")

    wb_bytes = io.BytesIO()
    wb.save(wb_bytes)
    wb_bytes.seek(0)
    return wb_bytes


@carga_masiva_bp.route('/partidos', methods=['GET', 'POST'])
@login_required
def partidos():
    series  = Serie.query.filter_by(activa=True).order_by(Serie.nombre).all()
    equipos = Equipo.query.filter_by(activo=True).order_by(Equipo.nombre).all()

    if request.method == 'POST':
        archivo = request.files.get('archivo')
        serie_id_default = request.form.get('serie_id', type=int)

        if not archivo or archivo.filename == '':
            flash('Debes seleccionar un archivo Excel.', 'danger')
            return redirect(url_for('carga_masiva.partidos'))

        if not archivo.filename.endswith(('.xlsx', '.xls')):
            flash('El archivo debe ser Excel (.xlsx).', 'danger')
            return redirect(url_for('carga_masiva.partidos'))

        try:
            from openpyxl import load_workbook
            wb = load_workbook(archivo, data_only=True)

            if 'partidos' not in wb.sheetnames:
                flash('El archivo no contiene la hoja "partidos". Usa la plantilla oficial.', 'danger')
                return redirect(url_for('carga_masiva.partidos'))

            ws = wb['partidos']

            # Precargar lookup dicts para búsqueda O(1)
            equipos_map = {e.codigo.upper(): e for e in Equipo.query.filter_by(activo=True).all()}
            series_map  = {s.codigo.upper(): s for s in Serie.query.filter_by(activa=True).all()}

            errores   = []
            partidos_validos = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Saltar filas completamente vacías
                if not any(row):
                    continue

                fecha_raw, hora_raw, serie_cod, local_cod, visita_cod, cancha, arbitro = (
                    (row + (None,) * 7)[:7]
                )

                # ── Validar fecha ──────────────────────────────────
                fecha = None
                if fecha_raw is None:
                    errores.append(f"Fila {row_num}: Fecha vacía.")
                elif isinstance(fecha_raw, (date, datetime)):
                    fecha = fecha_raw.date() if isinstance(fecha_raw, datetime) else fecha_raw
                else:
                    try:
                        fecha = datetime.strptime(str(fecha_raw).strip(), "%d/%m/%Y").date()
                    except ValueError:
                        errores.append(f"Fila {row_num}: Fecha «{fecha_raw}» inválida (usa DD/MM/AAAA).")

                # ── Validar hora ───────────────────────────────────
                hora = None
                if hora_raw is not None:
                    if isinstance(hora_raw, time):
                        hora = hora_raw
                    elif isinstance(hora_raw, datetime):
                        hora = hora_raw.time()
                    else:
                        try:
                            h, m = str(hora_raw).strip().split(':')
                            hora = time(int(h), int(m))
                        except Exception:
                            errores.append(f"Fila {row_num}: Hora «{hora_raw}» inválida (usa HH:MM).")

                # ── Validar serie ──────────────────────────────────
                serie = None
                if not serie_cod:
                    if serie_id_default:
                        serie = Serie.query.get(serie_id_default)
                    else:
                        errores.append(f"Fila {row_num}: Código de serie vacío.")
                else:
                    serie = series_map.get(str(serie_cod).strip().upper())
                    if not serie:
                        errores.append(f"Fila {row_num}: Serie «{serie_cod}» no existe en la BD.")

                # ── Validar equipos ────────────────────────────────
                local = visita = None
                if not local_cod:
                    errores.append(f"Fila {row_num}: Código de equipo local vacío.")
                else:
                    local = equipos_map.get(str(local_cod).strip().upper())
                    if not local:
                        errores.append(f"Fila {row_num}: Equipo local «{local_cod}» no existe en la BD.")

                if not visita_cod:
                    errores.append(f"Fila {row_num}: Código de equipo visita vacío.")
                else:
                    visita = equipos_map.get(str(visita_cod).strip().upper())
                    if not visita:
                        errores.append(f"Fila {row_num}: Equipo visita «{visita_cod}» no existe en la BD.")

                if local and visita and local.id == visita.id:
                    errores.append(f"Fila {row_num}: Local y visita son el mismo equipo ({local_cod}).")

                # Si la fila es válida hasta aquí, la acumulo
                if fecha and serie and local and visita and local != visita:
                    partidos_validos.append(dict(
                        fecha=fecha, hora=hora,
                        serie_id=serie.id, local_id=local.id, visita_id=visita.id,
                        cancha=str(cancha).strip() if cancha else None,
                        arbitro=str(arbitro).strip() if arbitro else None,
                    ))

            # ── Si hay cualquier error → rechazar toda la carga ───
            if errores:
                return render_template('carga_masiva/partidos.html',
                    series=series, equipos=equipos,
                    errores=errores, archivo_nombre=archivo.filename)

            if not partidos_validos:
                flash('El archivo no contiene filas con datos.', 'warning')
                return redirect(url_for('carga_masiva.partidos'))

            # ── Todo OK → insertar en BD ───────────────────────────
            for p in partidos_validos:
                db.session.add(Partido(estado='programado', **p))
            db.session.commit()

            flash(f'✅ {len(partidos_validos)} partido(s) cargados correctamente.', 'success')
            return redirect(url_for('partidos.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al procesar el archivo: {str(e)}', 'danger')
            return redirect(url_for('carga_masiva.partidos'))

    return render_template('carga_masiva/partidos.html',
        series=series, equipos=equipos, errores=[], archivo_nombre=None)


@carga_masiva_bp.route('/partidos/plantilla')
@login_required
def descargar_plantilla():
    """Descarga el Excel con los códigos reales de la BD."""
    excel = _generar_plantilla_con_datos()
    return send_file(
        excel,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_partidos.xlsx'
    )
